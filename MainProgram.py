#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from keras.models import load_model
from PIL import Image
import PIL.ImageOps 
import numpy as np
import matplotlib.image as mpimg 
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook


# In[40]:


#sorting the top three marks based on indexes
def Sorting():
    temp=0
    temp1=0
    for i in range(5):
        for j in range(5-i-1):
            if(sums[j] < sums[j+1]):
                temp=sums[j]
                sums[j]=sums[j+1]
                sums[j+1]=temp
                temp1=indexs[j]
                indexs[j]=indexs[j+1]
                indexs[j+1]=temp1

#Global Variables
sums=[0,0,0,0,0] #Total Sum of each question
indexs=[1,2,3,4,5] #Mapping the questions
Matrix = [[0 for x in range(4)] for y in range(5)] #Matrix of 5x4


img = Image.open("C:/mnist/Model/Ourimages/marks1.jpeg") #Image from Camera
model = load_model('mnistCNN.h5') #loading weights
new_image = img.resize((224, 280))  #resizing image to 224x280

#(x1,y1)-topLeft and (x2,y2)-bottomRight points of a perticular block
y1=5
y2=51
for i in range(5):
    x1=5
    x2=51
    for j in range(4): 
        area = (x1,y1,x2,y2)
        cropped_img = new_image.crop(area) #getting the image based on the dimenstions
       
        #Converting image to mnist format
        gray = cropped_img.convert('L') 
        bw = gray.point(lambda x: 0 if x<128 else 255, '1')
        bw.save(r"C:\mnist\Model\Ourimages\result_bw5.jpeg")
        image = Image.open(r'C:\mnist\Model\Ourimages\result_bw5.jpeg')
        inverted_image = PIL.ImageOps.invert(image)
        
        #Predicting the inverted_image(mnist formated image)
        result=[0,0,0,0,0,0,0,0,0,0] 
        for index in range(10):
            img = inverted_image.convert("L") #converting the image to gray scale
            img = img.resize((28,28)) #resizing the image to 28x28 pixels
            im2arr = np.array(img) #converting the image to numpy array
            im2arr = im2arr.reshape(1,28,28,1) #again reshaping the array
            y_pred = model.predict(im2arr) #predicting the number detected in the cell
            
            #storing the probability of detected number
            for idx in range(10):
                if y_pred[0,idx]>=1.0:
                    result[idx]+=1
                    
            max=result[0]
            residx=0
            
            #checking the max count of each number of result array
            for k in range(10):
                if(max<result[k]):
                    max=result[k]
                    residx=k
                    
        #Storing the recognized element in Matrix    
        if(residx>5):
            Matrix[i][j]=0
            sums[i]=sums[i]+Matrix[i][j]
        else:  
            Matrix[i][j]=residx
            sums[i]=sums[i]+Matrix[i][j]
    
        x1+=56
        x2+=56 #iterating through the next cell same row
    y1+=56
    y2+=56 #iterating through the next cell next row
    
#print (Matrix)
#print (sums)
Sorting()
#print (sums)
#print (indexs)

#Updating in the excel sheet
rowValue=10
wb = load_workbook(filename = 'Excel.xlsx')
sheet = wb.get_active_sheet()
for i in range(3):
    for j in range(4):
        obj = sheet.cell(row= rowValue  , column = ((indexs[i]*4)-1+j)) 
        obj.value = Matrix[indexs[i]-1][j]
rowValue+=1
wb.save('Excel.xlsx')


# In[ ]:





# In[ ]:




